from decimal import Decimal
import json
from io import BytesIO
import stripe
from datetime import datetime, timedelta, date
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_http_methods
from datetime import date

from apps.core.models import Reservation, Tour

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:
    openpyxl = None

stripe.api_key = settings.STRIPE_SECRET_KEY


# ============================================================
# TOUR DETAIL PAGE (IMPORTANT) -> envoie disabled_ranges + reservation
# ============================================================
@login_required
def tour_detail(request, tour_id):
    tour = get_object_or_404(Tour, id=tour_id)

    # Compute promo price for template display.
    if tour.is_promotion and (tour.discount_percent or 0) > 0:
        try:
            discount = (Decimal(100) - Decimal(tour.discount_percent)) / Decimal(100)
            tour.promo_price = (Decimal(tour.price_per_night) * discount).quantize(Decimal('0.01'))
        except Exception:
            tour.promo_price = None
    else:
        tour.promo_price = None

    # ✅ Reservation du user courant sur ce tour (active)
    reservation = Reservation.objects.filter(
        user=request.user,
        tour=tour
    ).exclude(status__in=["cancelled"]).order_by("-created_at").first()

    # ✅ Bloquer uniquement les ranges BOOKED par d'autres users
    booked_by_others = Reservation.objects.filter(
        tour=tour,
        status="booked",
    ).exclude(user=request.user)

    disabled_ranges = []
    booked_months = set()
    for r in booked_by_others:
        booked_months.add((r.start_date.year, r.start_date.month))

    # Disable entire months that have bookings
    for year, month in booked_months:
        # Get first and last day of the month
        from datetime import datetime
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)
        disabled_ranges.append({
            "from": first_day.strftime("%Y-%m-%d"),
            "to": last_day.strftime("%Y-%m-%d"),
        })

    return render(request, "booking.html", {
        "tour": tour,
        "reservation": reservation,
        "disabled_ranges": disabled_ranges,
        "STRIPE_PUBLIC_KEY": settings.STRIPE_PUBLIC_KEY,
        "activities_list": [activity.strip() for activity in tour.activities.split(',')] if tour.activities else [],
        "today": date.today(),
    })


# ============================================================
# CREATE BOOKING (ALWAYS PENDING)
# ============================================================
@login_required
def book_tour(request, tour_id):
    tour = get_object_or_404(Tour, id=tour_id)

    if request.method != "POST":
        return redirect("tour_detail", tour_id=tour.id)

    start = request.POST.get("start_date")
    end = request.POST.get("end_date")
    persons = int(request.POST.get("persons", 1))
    payment_method = request.POST.get("payment_method", "cash")

    try:
        start_date = datetime.strptime(start, "%Y-%m-%d").date()
        end_date = datetime.strptime(end, "%Y-%m-%d").date()
    except:
        messages.error(request, "Invalid date format")
        return redirect("tour_detail", tour_id=tour.id)

    nights = (end_date - start_date).days
    if nights <= 0:
        messages.error(request, "Invalid date range")
        return redirect("tour_detail", tour_id=tour.id)

    # ✅ conflit seulement avec BOOKED (pas pending)
    # Bloquer si une réservation existe dans le même mois
    conflict = Reservation.objects.filter(
        tour=tour,
        status="booked",
        start_date__year=start_date.year,
        start_date__month=start_date.month
    ).exclude(user=request.user).exists()

    if conflict:
        messages.error(request, "❌ This month is already fully booked. Please choose another month.")
        return redirect("tour_detail", tour_id=tour.id)

    # total
    base_price = Decimal(tour.price_per_night)

    if tour.is_promotion and tour.discount_percent > 0:
        discount = (Decimal(100) - Decimal(tour.discount_percent)) / Decimal(100)
        base_price = (base_price * discount).quantize(Decimal("0.01"))

    total = base_price * Decimal(nights) * Decimal(persons)

    # ✅ 10% discount for 5+ persons
    if persons >= 5:
        total = (total * Decimal("0.90")).quantize(Decimal("0.01"))

    # ✅ If user has pending/booked reservation, cancel it first (modification)
    existing = Reservation.objects.filter(
        user=request.user,
        tour=tour
    ).exclude(status__in=["cancelled", "rejected"]).first()

    if existing:
        existing.status = "cancelled"
        existing.save()
        messages.info(request, "ℹ️ Your previous booking has been updated with new dates.")

    Reservation.objects.create(
        user=request.user,
        tour=tour,
        start_date=start_date,
        end_date=end_date,
        num_persons=persons,
        total_price=total,

        booking_for_other=request.POST.get("booking_for") == "other",
        guest_full_name=request.POST.get("guest_full_name", ""),
        guest_phone=request.POST.get("guest_phone", ""),

        # ✅ ALWAYS pending
        status="pending",

        payment_method=payment_method,      # cash | transfer
        payment_status="unpaid",            # unpaid until paid
        stripe_payment_intent=""            # empty for now
    )

    messages.success(request, "✅ Booking request submitted. Waiting for admin validation.")
    return redirect("home")


# ============================================================
# ✅ Stripe Intent creation ONLY AFTER admin validation
# ============================================================
@login_required
@require_POST
def create_payment_intent_for_reservation(request, reservation_id):
    r = get_object_or_404(Reservation, id=reservation_id, user=request.user)

    # ✅ Only if booked + card + unpaid
    if r.status != "booked":
        return JsonResponse({"error": "Booking not validated yet."}, status=400)

    if r.payment_method != "card":
        return JsonResponse({"error": "This reservation is not card payment."}, status=400)

    if r.payment_status == "paid":
        return JsonResponse({"error": "Already paid."}, status=400)

    amount = int(float(r.total_price) * 100)

    intent = stripe.PaymentIntent.create(
        amount=amount,
        currency="mad",
        automatic_payment_methods={"enabled": True},
        metadata={
            "reservation_id": str(r.id),
            "tour_id": str(r.tour.id),
            "user": r.user.username,
        }
    )

    # ✅ store intent id (optional)
    r.stripe_payment_intent = intent.id
    r.save(update_fields=["stripe_payment_intent"])

    return JsonResponse({
        "client_secret": intent.client_secret,
        "intent_id": intent.id
    })


# ============================================================
# CANCEL reservation (ONLY if pending/rejected and still future)
# ============================================================
@login_required
def cancel_reservation(request, id):
    r = get_object_or_404(Reservation, id=id, user=request.user)

    today = date.today()
    # ✅ cannot cancel if booked OR date passed
    if r.status == "booked" or r.end_date < today:
        messages.info(request, "You cannot cancel this reservation.")
        return redirect("home")

    if r.status in ["pending", "rejected"]:
        r.status = "cancelled"
        r.save()
        messages.success(request, "✅ Reservation cancelled.")

    return redirect("home")


# ============================================================
# ADMIN
# ============================================================


def _get_profile_fields(user):
    try:
        profile = getattr(user, 'profile', None)
        if not profile:
            return None
        return {
            'profile_phone': getattr(profile, 'phone', ''),
            'profile_country': getattr(profile, 'country', ''),
            'profile_postal_code': getattr(profile, 'postal_code', ''),
        }
    except Exception:
        return None


def _reservation_report_headers():
    return [
        'reservation_id',
        'created_at',
        'status',
        'admin_note',
        'payment_method',
        'payment_status',
        'stripe_payment_intent',
        'tour_id',
        'tour_title',
        'tour_country',
        'destination',
        'start_date',
        'end_date',
        'nights',
        'num_persons',
        'total_price',
        'booking_for_other',
        'guest_full_name',
        'guest_phone',
        'user_id',
        'username',
        'email',
        'first_name',
        'last_name',
        'is_staff',
        'is_active',
        'date_joined',
        'last_login',
        'profile_phone',
        'profile_country',
        'profile_postal_code',
    ]


def _reservation_report_row(r):
    u = r.user
    t = r.tour
    destination_name = ''
    try:
        destination_name = getattr(getattr(t, 'destination', None), 'name', '') or ''
    except Exception:
        destination_name = ''

    profile_fields = _get_profile_fields(u) or {
        'profile_phone': '',
        'profile_country': '',
        'profile_postal_code': '',
    }

    return [
        r.id,
        r.created_at.isoformat() if getattr(r, 'created_at', None) else '',
        r.status,
        r.admin_note,
        r.payment_method,
        r.payment_status,
        r.stripe_payment_intent or '',
        getattr(t, 'id', ''),
        getattr(t, 'title', ''),
        getattr(t, 'country', ''),
        destination_name,
        r.start_date.isoformat() if r.start_date else '',
        r.end_date.isoformat() if r.end_date else '',
        getattr(r, 'nights', ''),
        r.num_persons,
        str(r.total_price),
        'yes' if r.booking_for_other else 'no',
        r.guest_full_name,
        r.guest_phone,
        u.id,
        u.username,
        u.email,
        u.first_name,
        u.last_name,
        'yes' if u.is_staff else 'no',
        'yes' if u.is_active else 'no',
        u.date_joined.isoformat() if u.date_joined else '',
        u.last_login.isoformat() if u.last_login else '',
        profile_fields['profile_phone'],
        profile_fields['profile_country'],
        profile_fields['profile_postal_code'],
    ]


def _reservations_to_xlsx_response(reservations_qs, filename: str) -> HttpResponse:
    if openpyxl is None:
        # Fallback: plain text if openpyxl isn't available (shouldn't happen in prod)
        response = HttpResponse('Excel export requires openpyxl.', content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.txt"'
        return response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reservations'

    headers = _reservation_report_headers()
    ws.append(headers)

    for r in reservations_qs:
        ws.append(_reservation_report_row(r))

    # Styling
    header_fill = PatternFill('solid', fgColor='1F2937')
    header_font = Font(color='FFFFFF', bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Add Excel table style (clear borders + banded rows)
    table = Table(displayName='ReservationsTable', ref=ws.auto_filter.ref)
    style = TableStyleInfo(
        name='TableStyleMedium9',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Best-effort column widths
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, min(ws.max_row, 250) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 45)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def admin_reservations(request):
    if not request.user.is_staff:
        return redirect("home")

    reservations_qs = Reservation.objects.select_related(
        'user',
        'tour',
        'tour__destination',
    ).all().order_by("-created_at")

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        selected_ids = request.POST.getlist('reservation_ids')
        selected_qs = reservations_qs
        if selected_ids:
            selected_qs = reservations_qs.filter(id__in=selected_ids)

        if action == 'delete_selected':
            if not selected_ids:
                messages.error(request, 'Select at least one reservation to delete.')
                return redirect('admin_reservations')

            deleted_count, _ = selected_qs.delete()
            messages.success(request, f'✅ Deleted {deleted_count} reservation(s).')
            return redirect('admin_reservations')

        if action == 'download_report':
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'reservations_report_{ts}.xlsx'
            return _reservations_to_xlsx_response(selected_qs, filename)

        messages.info(request, 'Choose an action to apply.')
        return redirect('admin_reservations')

    return render(request, "admin_reservations.html", {
        "reservations": reservations_qs,
        "reservations_count": reservations_qs.count(),
    })


@login_required
@require_POST
def delete_reservation(request, id):
    if not request.user.is_staff:
        return redirect("home")

    r = get_object_or_404(Reservation, id=id)
    r.delete()
    messages.success(request, '✅ Reservation deleted.')
    return redirect('admin_reservations')


@login_required
@require_http_methods(['GET'])
def download_reservation_report(request, id):
    if not request.user.is_staff:
        return redirect("home")

    qs = Reservation.objects.select_related('user', 'tour', 'tour__destination').filter(id=id)
    r = get_object_or_404(qs, id=id)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'reservation_{r.id}_report_{ts}.xlsx'
    return _reservations_to_xlsx_response(qs, filename)


@login_required
def validate_reservation(request, id):
    if not request.user.is_staff:
        return redirect("home")

    r = get_object_or_404(Reservation, id=id)
    r.status = "booked"
    r.save()
    messages.success(request, "✅ Reservation validated.")
    return redirect("admin_reservations")


@login_required
def reject_reservation(request, id):
    if not request.user.is_staff:
        return redirect("home")

    r = get_object_or_404(Reservation, id=id)
    r.status = "rejected"
    r.save()
    messages.error(request, "❌ Reservation rejected.")
    return redirect("admin_reservations")


@login_required
def mark_paid_reservation(request, id):
    if not request.user.is_staff:
        return redirect("home")

    r = get_object_or_404(Reservation, id=id)

    # ✅ allow manual marking for cash and transfer methods
    if (r.payment_method in ["cash", "transfer"]) and r.payment_status != "paid":
        r.payment_status = "paid"
        r.save()
        messages.success(request, "✅ Payment marked as PAID.")
    else:
        messages.info(request, "Nothing to update.")

    return redirect("admin_reservations")


# ============================================================
# STRIPE WEBHOOK -> update ONLY payment_status
# ============================================================
@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    try:
        event = stripe.Event.construct_from(json.loads(payload), stripe.api_key)
    except Exception:
        return JsonResponse({"status": "invalid"}, status=400)

    if event.type == "payment_intent.succeeded":
        intent = event.data.object

        # ✅ mark paid only
        Reservation.objects.filter(stripe_payment_intent=intent.id).update(
            payment_status="paid"
        )

    return JsonResponse({"status": "ok"})
@login_required
def admin_cancel_reservation(request, id):
    if not request.user.is_staff:
        return redirect("home")

    r = get_object_or_404(Reservation, id=id)

    # ✅ Admin can cancel ONLY booked + unpaid
    if r.status == "booked" and r.payment_status != "paid":
        r.status = "cancelled"
        r.save()
        messages.success(request, "✅ Reservation cancelled by admin.")
    else:
        messages.info(request, "Cannot cancel this reservation.")

    return redirect("admin_reservations")
